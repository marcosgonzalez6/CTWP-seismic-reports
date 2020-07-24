import tkinter as tk
from tkinter import filedialog
from PIL import Image, ImageTk
from openpyxl import load_workbook
from src.update_report import SeismicReport


class ReportDisplay(tk.Tk):
    def __init__(self):
        super(ReportDisplay, self).__init__()
        self.sr = SeismicReport()
        self.geometry("700x500")
        self.title("CTWP Seismic Reports")
        self.canvas = tk.Canvas(self, width=700, height=500)
        self.map_image = Image.open("caribbean_map3.png")
        self.map_image_copy = self.map_image.copy()
        self.background_image = ImageTk.PhotoImage(self.map_image)
        self.background = tk.Label(self.canvas, image=self.background_image)
        self.background.pack(expand='yes', fill='both')
        self.background.bind('<Configure>', self._resize_image)
        v2 = tk.StringVar(self)
        v2.set('Enter the year: ')
        self.year_entry = tk.Entry(master=self.canvas, textvariable=v2, width=20, justify='center')
        months = ['January', 'February', 'March', 'April', 'May', 'June',
                  'July', 'August', 'September', 'November', 'December']
        self.v1 = tk.StringVar(self)
        self.v1.set('--Select the month--')
        self.month_optionmenu = tk.OptionMenu(self.canvas, self.v1, *months)
        self.buttons()
        self.canvas.pack(expand='yes', fill='both')

    # source: https://stackoverflow.com/questions/24061099/tkinter-resize-background-image-to-window-size
    def _resize_image(self, event):
        new_width = event.width
        new_height = event.height
        self.map_image = self.map_image_copy.resize((new_width, new_height))
        self.background_image = ImageTk.PhotoImage(self.map_image)
        self.background.configure(image=self.background_image)

    def get_report(self):
        self.sr.report = load_workbook(filename=filedialog.askopenfilename())
        self.sr.clean_report()

    def prsn(self):
        try:
            data = load_workbook(filename=filedialog.askopenfilename()).active
            self.sr.update_prsn(data)
        except FileNotFoundError:
            print("PRSN file not found")

    def iris(self):
        try:
            data = load_workbook(filename=filedialog.askopenfilename()).active
            self.sr.update_iris(data)
        except FileNotFoundError:
            print("IRIS file not found")

    def ntwc(self):
        self.sr.month = self.v1.get()
        try:
            data = load_workbook(filename=filedialog.askopenfilename()).active
            self.sr.update_ntwc(data)
        except FileNotFoundError:
            print("NTWC file not found")

    def ptwc(self):
        try:
            data = load_workbook(filename=filedialog.askopenfilename()).active
            self.sr.update_ptwc(data)
        except FileNotFoundError:
            print("PTWC file not found")

    def complete_report(self):
        self.sr.year = self.year_entry.get()[-4:]
        self.sr.update_status()
        self.sr.analysis()
        self.sr.save()

    def buttons(self):
        report_button = tk.Button(self.canvas, text="Select last month's report", command=self.get_report, font=('helvetica', 12, 'bold'), width=25)
        prsn_button = tk.Button(self.canvas, text='Select PRSN data', command=self.prsn, font=('helvetica', 12, 'bold'), width=25)
        iris_button = tk.Button(self.canvas, text='Select IRIS data', command=self.iris, font=('helvetica', 12, 'bold'), width=25)
        ntwc_button = tk.Button(self.canvas, text='Select NTWC data', command=self.ntwc, font=('helvetica', 12, 'bold'), width=25)
        ptwc_button = tk.Button(self.canvas, text='Select PTWC data', command=self.ptwc, font=('helvetica', 12, 'bold'), width=25)
        complete_button = tk.Button(self.canvas, text='Finish report', command=self.complete_report, font=('helvetica', 12, 'bold'), width=25)
        self.canvas.create_window(500, 125, window=report_button)
        self.canvas.create_window(500, 175, window=prsn_button)
        self.canvas.create_window(500, 225, window=iris_button)
        self.canvas.create_window(500, 275, window=ntwc_button)
        self.canvas.create_window(500, 325, window=ptwc_button)
        self.canvas.create_window(500, 375, window=complete_button)

        self.canvas.create_window(200, 200, window=self.year_entry)
        self.canvas.create_window(200, 250, window=self.month_optionmenu)

    def entries(self):
        months = ['January', 'February', 'March', 'April', 'May', 'June',
                  'July', 'August', 'September', 'November', 'December']
        v1 = tk.StringVar(self)
        v1.set('--Select the month--')
        month_dropdown = tk.OptionMenu(self.canvas, v1, *months)
        self.sr.month = v1.get()

        self.canvas.create_window(150, 200, window=self.year_entry)
        self.canvas.create_window(150, 250, window=month_dropdown)


# TODO is global needed that many times??
#
# TODO add underscore to the beginning of each (internal) function
if __name__ == '__main__':
    root = ReportDisplay()
    root.mainloop()
