from openpyxl.styles import PatternFill, Font, Border, Alignment
from openpyxl import load_workbook
from copy import copy
import numpy as np


class SeismicReport:
    def __init__(self):
        self.month = None
        self.year = None
        self.report = None
        # try:
        #     self.report = load_workbook(file_name)
        #     self.clean_report()
        # except FileNotFoundError:
        #     print("Report file not found")

    def update_prsn(self, data):
        report_sheet = self.report.worksheets[0]
        i = 2  # counter to keep track of row number in report
        while report_sheet['F' + str(i)].value:
            new_latency = -1  # necessary because different sensors can report different latencies so we get the max
            for j in range(2, data.max_row + 1):
                if report_sheet['F' + str(i)].value.strip() == data['A' + str(j)].value.strip():
                    if new_latency < float(data['D' + str(j)].value):
                        new_latency = float(data['D' + str(j)].value)
                    if not report_sheet['H' + str(i)].value:  # column H is PRSN channel
                        print(f"PRSN channel added in line {i}")
                    elif data['B' + str(j)].value.strip() not in report_sheet['H' + str(i)].value.split('/'):
                        print(f"PRSN channel change in line {i}")
                    if report_sheet['G' + str(i)].value.strip() != data['C' + str(j)].value.strip():  # column G is network code
                        print(f"PRSN network code change in line {i}")
                    if j == data.max_row or report_sheet['F' + str(i)].value != data['A' + str(j + 1)].value:  # to stop search
                        break
            if new_latency > -1:
                self.update_latency(old_cell=report_sheet['M' + str(i)], new_cell_value=new_latency, agency="PRSN")
            elif report_sheet['M' + str(i)].value is not None:
                self.update_latency(old_cell=report_sheet['M' + str(i)], new_cell_value=None, agency="PRSN")
            i += 1

    def update_iris(self, data):
        report_sheet = self.report.worksheets[0]
        i = 2  # counter to keep track of row number in report
        while report_sheet['F' + str(i)].value:
            new_latency = -1  # necessary because different sensors can report different latencies so we get the max
            for j in range(1, data.max_row + 1):
                if data['A' + str(j)].value.strip() == "-Channel" \
                        and report_sheet['F' + str(i)].value.strip() == data['C' + str(j)].value.strip():
                    if new_latency < float(data['I' + str(j)].value.rstrip("%")):
                        new_latency = float(data['I' + str(j)].value.rstrip("%"))
                    if not report_sheet['I' + str(i)].value:  # column I is IRIS channel
                        print(f"IRIS channel added in line {i}")
                    else:
                        d = data['E' + str(j)].value.strip()
                        r = report_sheet['I' + str(i)].value.split('/')
                        s = float(data['I' + str(j)].value.rstrip("%"))
                        if d not in r and s > 3:
                            print(f"IRIS channel change in line {i}")
                    if report_sheet['G' + str(i)].value.strip() != data['B' + str(j)].value.strip():  # column G is network code
                        print(f"IRIS network code change in line {i}")
                    if data['A' + str(j + 1)].value.strip() != "-Channel":  # to stop search
                        break
            if new_latency > -1:
                self.update_latency(old_cell=report_sheet['N' + str(i)], new_cell_value=new_latency, agency="IRIS")
            elif report_sheet['N' + str(i)].value is not None:
                self.update_latency(old_cell=report_sheet['N' + str(i)], new_cell_value=None, agency="IRIS")
            i += 1

    def update_ntwc(self, data):
        # to find the column number of the current month
        for i in range(15, 27):
            c = data.cell(row=6, column=i)
            if c.value == self.month:
                month_column = c.column_letter

        # to find the minimum and maximum row in NTWC data with station information
        min_row = data.max_row
        for n in range(1, data.max_row + 1):
            if data['L' + str(n)].value == 'Station':
                min_row = n + 3  # station names start three rows after the cell with the title of the "Station" column
            if n > min_row and data['L' + str(n + 1)].value is None:
                max_row = n
                break

        report_sheet = self.report.worksheets[0]
        i = 2  # counter to keep track of row number in report
        while report_sheet['F' + str(i)].value:
            for j in range(min_row, max_row + 1):  # counter to keep track of row number in data
                if report_sheet['F' + str(i)].value.strip() == data['L' + str(j)].value.strip():
                    self.update_latency(report_sheet['O' + str(i)], 100 - float(data[month_column + str(j)].value), "NTWC")
                    # report['O' + str(i)].value = 100 - data[month_column + str(j)].value  # NTWC data is in column O
                    if report_sheet['J' + str(i)].value is None:  # column J is NTWC channel
                        print(f"NTWC channel added in line {i}")
                    elif report_sheet['J' + str(i)].value.strip() != data['M' + str(j)].value.strip():
                        print(f"NTWC channel change in line {i}")
                    if report_sheet['G' + str(i)].value.strip() != data['N' + str(j)].value.strip():  # column N is network code
                        print(f"NTWC network code change in line {i}")
                    break
                if j == max_row and report_sheet['O' + str(i)].value is not None:
                    self.update_latency(old_cell=report_sheet['O' + str(i)], new_cell_value=None, agency="NTWC")
            i += 1

    def update_ptwc(self, data):
        report_sheet = self.report.worksheets[0]
        i = 2  # counter to keep track of row number in report
        while report_sheet['F' + str(i)].value:
            for j in range(2, data.max_row + 1):  # counter to keep track of row number in data
                station_info = data['C' + str(j)].value.split('_')
                if report_sheet['F' + str(i)].value.strip() == station_info[0]:  # [0] is station code
                    self.update_latency(report_sheet['P' + str(i)], float(data['F' + str(j)].value), "PTWC")
                    # report['P' + str(i)].value = float(data['F' + str(j)].value)  # PTWC data is in column P
                    if report_sheet['K' + str(i)].value is None:  # column K is PTWC channel
                        print(f"PTWC channel added in line {i}")
                    elif report_sheet['K' + str(i)].value.strip() != station_info[1].split('.')[0]:
                        print(f"PTWC channel change in line {i}")
                    if report_sheet['G' + str(i)].value.strip() != station_info[1].split('.')[1]:
                        print(f"PTWC network code change in line {i}")
                    # to test if coordinates change by more than 1 degree
                    if report_sheet['D' + str(i)].value is None:
                        print(f"PTWC latitude missing in line {i}")
                    elif abs(float(report_sheet['D' + str(i)].value) - float(data['D' + str(j)].value)) > 1:
                        print(f"PTWC latitude change in line {i}")
                    if report_sheet['E' + str(i)].value is None:
                        print(f"PTWC longitude missing in line {i}")
                    elif abs(float(report_sheet['E' + str(i)].value) - float(data['E' + str(j)].value)) > 1:
                        print(f"PTWC longitude change in line {i}")
                    break
                if j == data.max_row and report_sheet['P' + str(i)].value is not None:
                    self.update_latency(old_cell=report_sheet['P' + str(i)], new_cell_value=None, agency="PTWC")
            i += 1

    def update_latency(self, old_cell, new_cell_value, agency):
        report_sheet = self.report.worksheets[0]
        change = None
        if old_cell.value is None:  # mark if an agency starts to report data for a station
            change = "(A)"
            print(f"Station {report_sheet['F' + str(old_cell.row)].value} added for {agency}")
        elif new_cell_value is None:
            change = "(X)"
            print(f"Station {report_sheet['F' + str(old_cell.row)].value} removed for {agency}")
        elif new_cell_value - float(old_cell.value) >= 10:  # mark any change of 10% or more
            change = "(U)"
        elif float(old_cell.value) - new_cell_value >= 10:
            change = "(D)"

        if change:
            comment = report_sheet['Q' + str(old_cell.row)]  # comments are in column Q
            if comment.value is None:
                comment.value = agency + " " + change + "; "
            elif change in report_sheet['Q' + str(old_cell.row)].value:
                n = comment.value.find(change)
                comment.value = comment.value[:n - 1] + ", " + agency + " " + comment.value[n:]
            else:
                comment.value += agency + " " + change + "; "

        old_cell.value = new_cell_value  # here is when the latency is actually updated

    def update_status(self):
        report_sheet = self.report.worksheets[0]
        i = 2  # counter to keep track of row number in report_sheet
        while report_sheet['L' + str(i)].value:
            contributing, down, multiple_agencies = False, False, False
            for c in range(13, 17):  # get the data cells of the 4 facilities for each station
                if report_sheet.cell(row=i, column=c).value is not None and report_sheet.cell(row=i, column=c).value >= 3.0:
                    if contributing or down:
                        multiple_agencies = True
                    contributing = True
                    down = False
                elif report_sheet.cell(row=i, column=c).value is not None and report_sheet.cell(row=i, column=c).value < 3.0:
                    if contributing or down:
                        multiple_agencies = True
                    down = True
            if contributing and report_sheet['L' + str(i)].value.strip() != "Contributing-RTX":
                report_sheet['L' + str(i)].value = "Contributing-RTX"
                report_sheet['L' + str(i)].font = Font(b=True)
                report_sheet['L' + str(i)].fill = PatternFill(start_color='CC99FF', end_color='CC99FF', fill_type='solid')
                print(f'Status change in line {i}')
            elif down and not contributing and report_sheet['L' + str(i)].value.strip() != "Down":
                report_sheet['L' + str(i)].value = "Down"
                report_sheet['L' + str(i)].font = Font(b=True)
                report_sheet['L' + str(i)].fill = PatternFill(start_color='CC99FF', end_color='CC99FF', fill_type='solid')
                print(f'Status change in line {i}')
            elif not contributing and not down:  # if no agency has data for this station
                report_sheet['F' + str(i)].fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
            elif not multiple_agencies:  # if only one agency has data for this station
                report_sheet['F' + str(i)].fill = PatternFill(start_color='FF99FF', end_color='FF99FF', fill_type='solid')
            if report_sheet['Q' + str(i)].value is not None:  # to remove trailing semicolon from comment column
                report_sheet['Q' + str(i)].value = report_sheet['Q' + str(i)].value.rstrip("; ")
            i += 1

    def analysis(self):
        # {key=Status: value=[column for Status, counter for 'ALL-status' sheet, counter for 'CARIBE-status' sheet]}
        status_columns = {'All': ['A', 0, 0], 'Contributing-RTX': ['B', 0, 0],
                          'Down': ['C', 0, 0], 'Existing': ['D', 0, 0], 'Gap': ['E', 0, 0],
                          'Unknown': ['F', 0, 0], 'Planned': ['G', 0, 0]}
        # [(column in '_-contributing' sheets, corresponding column in report)]
        contributing_columns = [('A', 'C'), ('B', 'F'), ('C', 'G'), ('D', 'L'), ('E', 'M'),
                                ('F', 'N'), ('G', 'O'), ('H', 'P'), ('I', 'Q')]
        i = 2  # counter to keep track of row number in report
        c = self.report.worksheets[0]['L' + str(i)]
        while c.value:
            if c.value in status_columns.keys():
                status_columns['All'][1] += 1
                self.report['ALL-status']['A' + str(status_columns['All'][1])].value = c.value
                self.report['ALL-status']['A' + str(status_columns['All'][1])].fill = copy(c.fill)
                status_columns[c.value][1] += 1
                self.report['ALL-status'][status_columns[c.value][0] + str(status_columns[c.value][1])].value = c.value
                self.report['ALL-status'][status_columns[c.value][0] + str(status_columns[c.value][1])].fill = copy(c.fill)
                if c.value == 'Contributing-RTX':
                    for x, y in contributing_columns:
                        self.report['ALL-contributing'][x + str(status_columns[c.value][1]+1)].value = \
                            self.report.worksheets[0][y + str(i)].value
                        self.report['ALL-contributing'][x + str(status_columns[c.value][1]+1)].fill = \
                            copy(self.report.worksheets[0][y + str(i)].fill)
                if self.report.worksheets[0]['C' + str(i)].value == 'CARIBE':
                    status_columns['All'][2] += 1
                    self.report['CARIBE-status']['A' + str(status_columns['All'][2])].value = c.value
                    self.report['CARIBE-status']['A' + str(status_columns['All'][2])].fill = copy(c.fill)
                    status_columns[c.value][2] += 1
                    self.report['CARIBE-status'][status_columns[c.value][0] + str(status_columns[c.value][2])].value = c.value
                    self.report['CARIBE-status'][status_columns[c.value][0] + str(status_columns[c.value][2])].fill = copy(c.fill)
                    if c.value == 'Contributing-RTX':
                        for x, y in contributing_columns:
                            self.report['CARIBE-contributing'][x + str(status_columns[c.value][2]+1)].value = \
                                self.report.worksheets[0][y + str(i)].value
                            self.report['CARIBE-contributing'][x + str(status_columns[c.value][2]+1)].fill = \
                                copy(self.report.worksheets[0][y + str(i)].fill)
            else:
                print(f"Incorrect status in line {i}")
            i += 1
            c = self.report.worksheets[0]['L' + str(i)]

        self.create_histograms(sheet_name='ALL-contributing')
        self.create_histograms(sheet_name='CARIBE-contributing')

    def create_histograms(self, sheet_name):
        # clear histograms
        sheet = self.report[sheet_name]
        for i in range(4, 32):
            if i in range(16, 20):
                continue
            sheet['L' + str(i)].value = None
            sheet['P' + str(i)].value = None

        # find last row with data
        n = 2
        while sheet['D' + str(n)].value:
            n += 1

        # {key=Agency: value=(column with Agency data, range of Agency histogram output, blanks in Agency)}
        hist_dic = {'PRSN': ('E', 'L4:L15', 0), 'IRIS': ('F', 'P4:P15', 0),
                    'NTWC': ('G', 'L20:L31', 0), 'PTWC': ('H', 'P20:P31', 0)}
        for k, v in hist_dic.items():
            # data_col is the list of all cells in the agency column with a reported latency
            data_col = [c[0].value for c in sheet[v[0] + '2:' + v[0] + str(n-1)] if c[0].value is not None]
            # hist is a list of length 10 with the quantities of each bin
            hist, _ = np.histogram(a=data_col, bins=10, range=(0, 100))
            for hist_cell, hist_result in zip(sheet[v[1]], hist):
                hist_cell[0].value = hist_result
            sheet[v[1]][10][0].value = 0    # the "More" bin
            sheet[v[1]][11][0].value = n - len(data_col) - 2    # counting the blanks

    def clean_report(self):  # report must be a Workbook, not a Worksheet
        a = 2  # counter to keep track of row number in main sheet of report
        while self.report.worksheets[0]['F' + str(a)].value:
            for c in self.report.worksheets[0]['A' + str(a) + ':L' + str(a)][0]:
                c.fill = PatternFill(fill_type=None)  # clear cells with color filling
            if self.report.worksheets[0]['Q' + str(a)].value:
                self.report.worksheets[0]['Q' + str(a)].value = None  # clear comments column
            a += 1

        b = 1  # counter to keep track of row number in ALL-status sheet of report
        while self.report['ALL-status']['A' + str(b)].value:
            for c in self.report['ALL-status']['A' + str(b) + ':I' + str(b)][0]:
                c.fill = PatternFill(fill_type=None)
                c.border = Border(outline=False)
                c.alignment = Alignment(horizontal='center')
                c.font = Font(b=True)
                c.value = None
            b += 1

        d = 1  # counter to keep track of row number in CARIBE-status sheet of report
        while self.report['CARIBE-status']['A' + str(d)].value:
            for c in self.report['CARIBE-status']['A' + str(d) + ':I' + str(d)][0]:
                c.fill = PatternFill(fill_type=None)
                c.border = Border(outline=False)
                c.alignment = Alignment(horizontal='center')
                c.font = Font(b=True)
                c.value = None
            d += 1

        e = 2  # counter to keep track of row number in ALL-contributing sheet of report
        while self.report['ALL-contributing']['A' + str(e)].value:
            for c in self.report['ALL-contributing']['A' + str(e) + ':I' + str(e)][0]:
                c.fill = PatternFill(fill_type=None)
                c.border = Border(outline=False)
                c.value = None
            e += 1

        f = 2  # counter to keep track of row number in CARIBE-contributing sheet of report
        while self.report['CARIBE-contributing']['A' + str(f)].value:
            for c in self.report['CARIBE-contributing']['A' + str(f) + ':I' + str(f)][0]:
                c.fill = PatternFill(fill_type=None)
                c.border = Border(outline=False)
                c.value = None
            f += 1

    def save(self):
        self.report.worksheets[0].title = f"{self.month[:3]}{self.year[2:]}"
        self.report.save(f"Seismic Report {self.month[:3]} {self.year}.xlsx")


def main():
    report = SeismicReport(file_name="SeismicDataAvailability_December2019.xlsx")
    ptwc_data = load_workbook(filename="Caribbean_stats_20200101-20200131.xlsx").active
    iris_data = load_workbook(filename="IRIS_Uptime_Report_for__CARIBE-EWS_-_2020_01_01-2020_01_31.xlsx").active
    prsn_data = load_workbook(filename="PRSN Jan 2020.xlsx").active
    ntwc_data = load_workbook(filename="Book2.xlsx").active  # get the first sheet
    report.update_prsn(prsn_data)
    report.update_iris(iris_data)
    report.update_ntwc(ntwc_data)
    report.update_ptwc(ptwc_data)
    report.update_status()
    report.analysis()
    report.save()


if __name__ == "__main__":
    main()
