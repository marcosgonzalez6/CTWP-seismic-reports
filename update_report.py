import sys
import tkinter as tk
from tkinter import filedialog
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font


def update_prsn(data, report):
    i = 2  # counter to keep track of row number in report
    while report['F' + str(i)].value:
        new_latency = -1  # necessary because different sensors can report different latencies so we get the max
        for j in range(2, data.max_row+1):
            if report['F' + str(i)].value.strip() == data['A' + str(j)].value.strip():
                if new_latency < float(data['D' + str(j)].value):
                    new_latency = float(data['D' + str(j)].value)
                if not report['H' + str(i)].value:  # column H is PRSN channel
                    print(f"PRSN channel added in line {i}")
                elif data['B' + str(j)].value.strip() not in report['H' + str(i)].value.split('/'):
                    print(f"PRSN channel change in line {i}")
                if report['G' + str(i)].value.strip() != data['C' + str(j)].value.strip():   # column G is network code
                    print(f"PRSN network code change in line {i}")
                if j == data.max_row or report['F' + str(i)].value != data['A' + str(j+1)].value:    # to stop search
                    break
        if new_latency > -1:
            update_latency(report['M' + str(i)], new_latency, "PRSN", report)
        elif report['M' + str(i)].value:
            update_latency(report['M' + str(i)], None, "PRSN", report)
        i += 1


def update_iris(data, report):
    i = 2  # counter to keep track of row number in report
    while report['F' + str(i)].value:
        new_latency = -1  # necessary because different sensors can report different latencies so we get the max
        for j in range(2, data.max_row+1):
            if data['A' + str(j)].value.strip() == "-Channel" \
                    and report['F' + str(i)].value.strip() == data['C' + str(j)].value.strip():
                if new_latency < float(data['I' + str(j)].value.rstrip("%")):
                    new_latency = float(data['I' + str(j)].value.rstrip("%"))
                if not report['I' + str(i)].value:   # column I is IRIS channel
                    print(f"IRIS channel added in line {i}")
                else:
                    d = data['E' + str(j)].value.strip()
                    r = report['I' + str(i)].value.split('/')
                    s = float(data['I' + str(j)].value.rstrip("%"))
                    if d not in r and s > 3:
                        print(f"IRIS channel change in line {i}")
                if report['G' + str(i)].value.strip() != data['B' + str(j)].value.strip():   # column G is network code
                    print(f"IRIS network code change in line {i}")
                if data['A' + str(j+1)].value.strip() != "-Channel":    # to stop search
                    break
        if new_latency > -1:
            update_latency(report['N' + str(i)], new_latency, "IRIS", report)
        elif report['N' + str(i)].value:
            update_latency(report['N' + str(i)], None, "IRIS", report)
        i += 1


def update_ntwc(data, report, month):
    # to find the column number of the current month
    for i in range(15, 27):
        c = data.cell(row=6, column=i)
        if c.value == month:
            month_column = c.column_letter

    # to find the minimum and maximum row in NTWC data with station information
    min_row = data.max_row
    for n in range(1, data.max_row):
        if data['L' + str(n)].value == 'Station':
            min_row = n+3
        if n > min_row and data['L' + str(n + 1)].value is None:
            max_row = n
            break

    i = 2  # counter to keep track of row number in report
    while report['F' + str(i)].value:
        for j in range(min_row, max_row):  # counter to keep track of row number in data
            if report['F' + str(i)].value.strip() == data['L' + str(j)].value.strip():
                update_latency(report['O' + str(i)], 100 - float(data[month_column + str(j)].value), "NTWC", report)
                # report['O' + str(i)].value = 100 - data[month_column + str(j)].value  # NTWC data is in column O
                if not report['J' + str(i)].value:  # column J is NTWC channel
                    print(f"NTWC channel added in line {i}")
                elif report['J' + str(i)].value.strip() != data['M' + str(j)].value.strip():
                    print(f"NTWC channel change in line {i}")
                if report['G' + str(i)].value.strip() != data['N' + str(j)].value.strip():   # column N is network code
                    print(f"NTWC network code change in line {i}")
                break
        if j == data.max_row and report['O' + str(i)].value:
            update_latency(report['O' + str(i)], None, "NTWC", report)
        i += 1


def update_ptwc(data, report):
    i = 2  # counter to keep track of row number in report
    while report['F' + str(i)].value:
        for j in range(2, data.max_row+1):  # counter to keep track of row number in data; UPDATE if number of stations changes
            station_info = data['C' + str(j)].value.split('_')
            if report['F' + str(i)].value.strip() == station_info[0]:   # [0] is station code
                update_latency(report['P' + str(i)], float(data['F' + str(j)].value), "PTWC", report)
                # report['P' + str(i)].value = float(data['F' + str(j)].value)  # PTWC data is in column P
                if not report['K' + str(i)].value:     # column K is PTWC channel
                    print(f"PTWC channel added in line {i}")
                elif report['K' + str(i)].value.strip() != station_info[1].split('.')[0]:
                    print(f"PTWC channel change in line {i}")
                if report['G' + str(i)].value.strip() != station_info[1].split('.')[1]:
                    print(f"PTWC network code change in line {i}")
                # to test if coordinates change by more than 1 degree
                if not report['D' + str(i)].value:
                    print(f"PWTC latitude missing in line {i}")
                elif abs(float(report['D' + str(i)].value) - float(data['D' + str(j)].value)) > 1:
                    print(f"PWTC latitude change in line {i}")
                if not report['E' + str(i)].value:
                    print(f"PWTC longitude missing in line {i}")
                elif abs(float(report['E' + str(i)].value) - float(data['E' + str(j)].value)) > 1:
                    print(f"PWTC longitude change in line {i}")
                break
        if j == data.max_row and report['P' + str(i)].value:
            update_latency(report['P' + str(i)], None, "PTWC", report)
        i += 1


def update_latency(old_cell, new_cell_value, agency, report):
    change = None
    if old_cell.value is None:      # mark if an agency starts to report data for a station
        change = "(A)"
        print(f"Station {report['F' + str(old_cell.row)].value} added for {agency}")
    elif new_cell_value is None:
        change = "(X)"
        print(f"Station {report['F' + str(old_cell.row)].value} removed for {agency}")
    elif new_cell_value - float(old_cell.value) >= 10:      # mark any change of 10% or more
        change = "(U)"
    elif float(old_cell.value) - new_cell_value >= 10:
        change = "(D)"

    if change:
        comment = report['Q' + str(old_cell.row)]       # comments are in column Q
        if comment.value is None:
            comment.value = agency + " " + change + "; "
        elif change in report['Q' + str(old_cell.row)].value:
            n = comment.value.find(change)
            comment.value = comment.value[:n-1] + ", " + agency + " " + comment.value[n:]
        else:
            comment.value += agency + " " + change + "; "

    old_cell.value = new_cell_value  # here is when the latency is actually updated


def update_status(report):
    i = 2  # counter to keep track of row number in report
    while report['L' + str(i)].value:
        contributing, down, multiple_agencies = False, False, False
        for c in range(13, 17):         # get the data cells of the 4 facilities for each station
            if report.cell(row=i, column=c).value is not None and report.cell(row=i, column=c).value >= 3.0:
                if contributing or down:
                    multiple_agencies = True
                contributing = True
                down = False
            elif report.cell(row=i, column=c).value is not None and report.cell(row=i, column=c).value < 3.0:
                if contributing or down:
                    multiple_agencies = True
                down = True
        if contributing and report['L' + str(i)].value.strip() != "Contributing-RTX":
            report['L' + str(i)].value = "Contributing-RTX"
            report['L' + str(i)].font = Font(b=True)
            report['L' + str(i)].fill = PatternFill(start_color='CC99FF', end_color='CC99FF', fill_type='solid')
            print(f'Status change in line {i}')
        elif down and report['L' + str(i)].value.strip() != "Down":
            report['L' + str(i)].value = "Down"
            report['L' + str(i)].font = Font(b=True)
            report['L' + str(i)].fill = PatternFill(start_color='CC99FF', end_color='CC99FF', fill_type='solid')
            print(f'Status change in line {i}')
        elif not contributing and not down:         # if no agency has data for this station
            report['F' + str(i)].fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
        elif not multiple_agencies:                 # if only one agency has data for this station
            report['F' + str(i)].fill = PatternFill(start_color='FF99FF', end_color='FF99FF', fill_type='solid')
        if report['Q' + str(i)].value is not None:  # to remove trailing semicolon from comment column
            report['Q' + str(i)].value = report['Q' + str(i)].value.rstrip("; ")
        i += 1


def clean_report(report):
    i = 2
    while report['F' + str(i)].value:
        for c in report['A' + str(i) + ':L' + str(i)][0]:
            c.fill = PatternFill(fill_type=None)    # clear cells with color filling
        if report['Q' + str(i)]:
            report['Q' + str(i)] = None             # clear comments column
        i += 1


def main():
    ptwc_data = load_workbook(filename="Caribbean_stats_20200101-20200131.xlsx").active
    iris_data = load_workbook(filename="IRIS_Uptime_Report_for__CARIBE-EWS_-_2020_01_01-2020_01_31.xlsx").active
    prsn_data = load_workbook(filename="PRSN Jan 2020.xlsx").active
    ntwc_data = load_workbook(filename="Book1.xlsx").active  # get the first sheet
    report = load_workbook(filename="SeismicDataAvailability_December2019.xlsx")
    report_sheet = report.active
    clean_report(report_sheet)
    source = "iris"  # FOR TESTING
    month = "January"  # FOR TESTING
    # if source == "prsn":
    update_prsn(prsn_data, report_sheet)
    # elif source == "iris":
    update_iris(iris_data, report_sheet)
    # elif source == "ntwc":
    update_ntwc(ntwc_data, report_sheet, month)
    # elif source == "ptwc":
    update_ptwc(ptwc_data, report_sheet)
    update_status(report_sheet)
    report.save("Seismic Report Jan 2020.xlsx")


def main_graphic():
    root = tk.Tk()
    root.title("CTWP Seismic Reports")
    canvas1 = tk.Canvas(root, width=500, height=500, bg='lightsteelblue')
    canvas1.pack()
    browseButton_Excel = tk.Button(text='Choose Excel with Seismic Data', command=getExcel, bg='green', fg='white',
                                   font=('helvetica', 12, 'bold'))
    canvas1.create_window(250, 250, window=browseButton_Excel)
    root.mainloop()


def getExcel():
    global df

    import_file_path = filedialog.askopenfilename()
    # df = pd.read_excel(import_file_path)
    df = load_workbook(filename=import_file_path).active
    print(df['A2'])


if __name__ == "__main__":
    main_graphic()
