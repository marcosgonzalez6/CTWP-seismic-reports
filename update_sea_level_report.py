from openpyxl import load_workbook
# datos de PTWC deben estar en un archivo excel en formato .xlsx
# Para hacer el cambio solo deben ....





sealevel = load_workbook('CTWP_Sea_Level_Report_November2020.xlsx')
ptwc_data = load_workbook('PTWC_Data.xlsx')

# Encuentra la ultima fila de la columna A con dato del excel Master
#-----------------------------------------------------
sheet = sealevel.active
last_row_master = sheet.max_row
while sheet.cell(column=1, row=last_row_master).value is None and last_row_master > 0:
    last_row_master -= 1
#last_col_a_value = sheet.cell(column=1, row=last_row).value
print(last_row_master)
#-----------------------------------------------------

# Encuentra la ultima fila de la columna A con dato del excel PTWC
#-----------------------------------------------------
sheet = ptwc_data.active
last_row_ptwc = sheet.max_row
while sheet.cell(column=1, row=last_row_ptwc).value is None and last_row_ptwc > 0:
    last_row_ptwc -= 1
#last_col_a_value = sheet.cell(column=1, row=last_row).value
print(last_row_ptwc)
#-----------------------------------------------------

# Codigo de VBA de excel
#-------------------------------------------------------------------------------------------------------
#For i = 1 To 272
#    For j = 2 To 187
#        If Worksheets("Report").Cells(i, 1).Value = Worksheets("PTWCData").Cells(j, 1).Value Then
#            If Worksheets("Report").Cells(i, 2).Value = Worksheets("PTWCData").Cells(j, 2).Value Then
#                Worksheets("PTWCData").Cells(j, 3).Copy
#                Worksheets("Report").Activate
#                Worksheets("Report").Cells(i, 3).Select
#                ActiveSheet.Paste
#            End If
#        End If
#    Next
#Next
#-------------------------------------------------------------------------------------------------------

# traduccion a Puthon
# Codigo compara los nombres de estacion y sensores en los archivos PTWC_Data y el master.
# si los valores igualan, los datos de PTWC para esta estacion y sensor son copiado y pegados
# en el Master para la respectiva estacion y sensor.
for i in range(9,last_row_master,1):
    for j in range(2,last_row_ptwc,1):
        master_a = sealevel.active.cell(row=i, column=3)
        ptwc_a = ptwc_data.active.cell(row=j, column=1)
        if  master_a.value == ptwc_a.value:
            master_b = sealevel.active.cell(row=i, column=4)
            ptwc_b = ptwc_data.active.cell(row=j, column=2)
            if master_b.value == ptwc_b.value:
                ptwc_data_value = ptwc_data.active.cell(row=j, column=3).value
                sealevel.active.cell(row=i, column=25).value = ptwc_data_value
sealevel.save('Test1.xlsx')
#--------------------------------------------------------------------------------------------------------



