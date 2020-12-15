import tkinter as tk
from tkinter import filedialog, ttk
from PIL import Image, ImageTk
from openpyxl import load_workbook
from update_report import SeismicReport
import PyInstaller


class ReportDisplay(tk.Tk):
    def __init__(self):
        super(ReportDisplay, self).__init__()
        self.sr = SeismicReport()
        self.geometry("700x500+200+100")
        self.title("CTWP Seismic Reports")
        self.canvas = tk.Canvas(self, width=700, height=500)

        # background image
        self.map_image = Image.open("images/caribbean_map.png")
        self.map_image_copy = self.map_image.copy()
        self.background_image = ImageTk.PhotoImage(self.map_image)
        self.background = tk.Label(self.canvas, image=self.background_image)
        self.background.pack(expand='yes', fill='both')
        self.background.bind('<Configure>', self._resize_image)

        # widgets on main window
        self.frame = tk.Frame(master=self.canvas)
        year_text = tk.StringVar(self)         # variable to store the year
        year_text.set('Enter the year: ')
        self.year_entry = tk.Entry(master=self.canvas, textvariable=year_text, justify='left', width=17)
        months = ['January', 'February', 'March', 'April', 'May', 'June',
                  'July', 'August', 'September', 'October', 'November', 'December']
        self.optionmenu_text = tk.StringVar(self)
        self.optionmenu_text.set('--Select the month--')
        self.month_optionmenu = tk.OptionMenu(self.canvas, self.optionmenu_text, *months)
        self.buttons()

        self.canvas.pack(expand='yes', fill='both')

    # source: https://stackoverflow.com/questions/24061099/tkinter-resize-background-image-to-window-size
    def _resize_image(self, event):
        new_width = event.width
        new_height = event.height
        self.map_image = self.map_image_copy.resize((new_width, new_height))
        self.background_image = ImageTk.PhotoImage(self.map_image)
        self.background.configure(image=self.background_image)

    def get_report(self):
        self.sr.report = load_workbook(filename=filedialog.askopenfilename(title="Select last month's report"))

    def prsn(self):
        try:
            data = load_workbook(filename=filedialog.askopenfilename(title="Select PRSN data")).active
            self.sr.prsn_data = data
        except FileNotFoundError:
            print("PRSN file not found")

    def iris(self):
        try:
            self.sr.iris_data = load_workbook(filename=filedialog.askopenfilename(title="Select IRIS data")).active
        except FileNotFoundError:
            print("IRIS file not found")

    def ntwc(self):
        try:
            self.sr.ntwc_data = load_workbook(filename=filedialog.askopenfilename(title="Select NTWC data")).active
        except FileNotFoundError:
            print("NTWC file not found")

    def ptwc(self):
        try:
            self.sr.ptwc_data = load_workbook(filename=filedialog.askopenfilename(title="Select PTWC data")).active
        except FileNotFoundError:
            print("PTWC file not found")

    def display_instructions(self):
        instructions_window = tk.Tk()
        instructions_window.wm_title("Instructions")
        container = ttk.Frame(instructions_window)
        ttk.Label(master=instructions_window, text=open('instructions.txt', 'r').read(), wraplength=500).pack()
        # container.pack()
        instructions_window.mainloop()

    # source: https://blog.tecladocode.com/tkinter-scrollable-frames/
    def create_output_log(self):
        output_log_window = tk.Tk()
        output_log_window.wm_title("Output Log")
        # output_log_window.geometry("700x500+200+100")
        container = ttk.Frame(output_log_window)
        canvas2 = tk.Canvas(container)
        scrollbar = ttk.Scrollbar(container, orient='vertical', command=canvas2.yview)
        scrollable_frame = ttk.Frame(canvas2)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas2.configure(
                scrollregion=canvas2.bbox("all")
            )
        )

        canvas2.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas2.configure(yscrollcommand=scrollbar.set)
        ttk.Label(master=scrollable_frame, text=self.sr.comments, width=500).pack()
        container.pack()
        canvas2.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        output_log_window.mainloop()

    def complete_report(self):
        self.sr.year = self.year_entry.get()[-4:]
        self.sr.month = self.optionmenu_text.get()
        self.sr.clear_report()
        self.sr.update_prsn()
        self.sr.update_iris()
        self.sr.update_ntwc()
        self.sr.update_ptwc()
        self.sr.update_status()
        self.sr.analysis()
        self.sr.save()
        output_log = tk.Button(self.canvas, text='Output log', command=self.create_output_log, font=('helvetica', 12, 'bold'))
        self.canvas.create_window(500, 400, window=output_log)

    def buttons(self):
        instructions_button = tk.Button(self.canvas, text="Instructions", command=self.display_instructions, font=('helvetica', 12, 'bold'), width=15)
        report_button = tk.Button(self.canvas, text="Select last month's report", command=self.get_report, font=('helvetica', 12, 'bold'), width=25)
        prsn_button = tk.Button(self.canvas, text='Select PRSN data', command=self.prsn, font=('helvetica', 12, 'bold'), width=25)
        iris_button = tk.Button(self.canvas, text='Select IRIS data', command=self.iris, font=('helvetica', 12, 'bold'), width=25)
        ntwc_button = tk.Button(self.canvas, text='Select NTWC data', command=self.ntwc, font=('helvetica', 12, 'bold'), width=25)
        ptwc_button = tk.Button(self.canvas, text='Select PTWC data', command=self.ptwc, font=('helvetica', 12, 'bold'), width=25)
        complete_button = tk.Button(self.canvas, text='Complete report', command=self.complete_report, font=('helvetica', 12, 'bold'), width=25)
        self.canvas.create_window(500, 125, window=report_button)
        self.canvas.create_window(500, 175, window=prsn_button)
        self.canvas.create_window(500, 225, window=iris_button)
        self.canvas.create_window(500, 275, window=ntwc_button)
        self.canvas.create_window(500, 325, window=ptwc_button)
        self.canvas.create_window(500, 375, window=complete_button)
        self.canvas.create_window(80, 40, window=instructions_button)

        self.canvas.create_window(200, 200, window=self.year_entry)
        self.canvas.create_window(200, 250, window=self.month_optionmenu)

    def entries(self):
        months = ['January', 'February', 'March', 'April', 'May', 'June',
                  'July', 'August', 'September', 'November', 'December']
        v1 = tk.StringVar(self)
        v1.set('--Select the month--')
        month_dropdown = tk.OptionMenu(self.canvas, v1, *months)
        self.sr.month = v1.get()

        self.canvas.create_window(150, 200, window=self.year_entry)
        self.canvas.create_window(150, 250, window=month_dropdown)


if __name__ == '__main__':
    root = ReportDisplay()
    root.mainloop()
